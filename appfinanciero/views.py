# appfinanciero/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from .models import Product, InventoryMove, Employee, PayrollPeriod, PayrollConfig, TaxBracket
from .forms import ProductForm, InventoryMoveForm, EmployeeForm, PayrollConfigForm, TaxBracketForm
from .services import valuacion_inventario, generar_planilla, _get_config
from .models import InventoryMove, PayrollConfig, TaxBracket, PayrollLine, PayrollPeriod, Employee

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import LETTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from decimal import Decimal, ROUND_HALF_UP
from reportlab.lib.styles import getSampleStyleSheet
from .forms import (
    ProductForm, InventoryMoveForm, EmployeeForm, PayrollConfigForm, TaxBracketForm,
    EntradaRapidaForm, SalidaRapidaForm
)
from django.utils import timezone            # ← FALTA ESTE
from decimal import Decimal  
import datetime
from decimal import Decimal

# appfinanciero/views.py (sustituir la función home existente)
from decimal import Decimal, ROUND_HALF_UP
from django.db import models   # si ya lo importaste arriba, no lo dupliques

def home(request):
    ultimos_productos = Product.objects.order_by('-id')[:8]
    ultimos_movimientos = InventoryMove.objects.order_by('-fecha')[:8]
    total_productos = Product.objects.count()
    total_empleados = Employee.objects.count()
    movimientos_30d = InventoryMove.objects.filter(fecha__gte=timezone.now()-timezone.timedelta(days=30)).count()

    # obtener el último periodo (orden descendente por año/mes)
    ultimo_periodo = PayrollPeriod.objects.order_by('-anio', '-mes').first()

    lineas_planilla = []
    total_liquido = Decimal('0.00')

    if ultimo_periodo:
        # generar planilla (si algo falla no rompe la página)
        try:
            generar_planilla(ultimo_periodo)
        except Exception as e:
            messages.error(request, f"Error al generar planilla: {e}")

        # traer primeras N líneas para mostrar en el dashboard
        qs = ultimo_periodo.lineas.select_related('empleado').order_by('empleado__nombre')[:6]
        for l in qs:
            # asegurar formato decimal a 2 decimales
            l.salario_base = (l.salario_base or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            l.isss = (l.isss or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            l.afp = (l.afp or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            l.renta = (l.renta or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            l.liquido = (l.liquido or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        lineas_planilla = list(qs)

        # suma total líquido del periodo
        agg = ultimo_periodo.lineas.aggregate(total=models.Sum('liquido'))
        total_liquido = (agg.get('total') or Decimal('0.00'))
        total_liquido = Decimal(total_liquido).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    return render(request, 'home.html', {
        'ultimos_productos': ultimos_productos,
        'ultimos_movimientos': ultimos_movimientos,
        'total_productos': total_productos,
        'total_empleados': total_empleados,
        'movimientos_30d': movimientos_30d,
        'ultimo_periodo': ultimo_periodo,
        'lineas_planilla': lineas_planilla,
        'total_liquido': total_liquido,
    })

# -------- Productos --------
def product_list(request):
    return render(request, 'productos/list.html', {'items': Product.objects.all()})

def product_create(request):
    form = ProductForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Producto creado.'); return redirect('product_list')
    return render(request, 'productos/form.html', {'form': form, 'titulo': 'Nuevo producto'})

def product_edit(request, pk):
    obj = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Producto actualizado.'); return redirect('product_list')
    return render(request, 'productos/form.html', {'form': form, 'titulo': 'Editar producto'})

@require_http_methods(['POST'])
def product_delete(request, pk):
    get_object_or_404(Product, pk=pk).delete()
    messages.info(request, 'Producto eliminado.')
    return redirect('product_list')

# -------- Movimientos --------
def move_list(request):
    return render(request, 'movimientos/list.html', {'items': InventoryMove.objects.select_related('producto').all()})

def move_create(request):
    form = InventoryMoveForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Movimiento registrado.'); return redirect('move_list')
    return render(request, 'movimientos/form.html', {'form': form, 'titulo': 'Nuevo movimiento'})

def reporte_inventario(request, producto_id, metodo):
    producto = get_object_or_404(Product, pk=producto_id)

    # ---- Captura rápida ----
    entrada_form = EntradaRapidaForm()
    salida_form  = SalidaRapidaForm()

    if request.method == 'POST':
        if 'add_entrada' in request.POST:
            entrada_form = EntradaRapidaForm(request.POST)
            if entrada_form.is_valid():
                cd = entrada_form.cleaned_data
                InventoryMove.objects.create(
                    producto=producto,
                    fecha=timezone.now(),
                    tipo=InventoryMove.IN,
                    cantidad=cd['cantidad'],
                    costo_unitario=cd['costo_unitario'],
                    referencia=cd.get('referencia', '')
                )
                messages.success(request, 'Entrada registrada.')
                return redirect('reporte_inventario', producto_id=producto.id, metodo=metodo)
        elif 'add_salida' in request.POST:
            salida_form = SalidaRapidaForm(request.POST)
            if salida_form.is_valid():
                cd = salida_form.cleaned_data
                try:
                    # crear y validar (en caso tengas la validación de stock en clean())
                    move = InventoryMove(
                        producto=producto,
                        fecha=timezone.now(),
                        tipo=InventoryMove.OUT,
                        cantidad=cd['cantidad'],
                        referencia=cd.get('referencia', '')
                    )
                    move.full_clean()
                    move.save()
                    messages.success(request, 'Salida registrada.')
                    return redirect('reporte_inventario', producto_id=producto.id, metodo=metodo)
                except ValidationError as e:
                    messages.error(request, '; '.join(sum(e.message_dict.values(), [])))
                except Exception as e:
                    messages.error(request, str(e))

    # ---- Valuación y armado de capas (con subtotal) ----
    try:
        existencia, costo_total, costo_promedio, capas = valuacion_inventario(producto, metodo.upper())
        capas_ui = []
        for c in capas:
            sub = c['qty'] * c['cost']
            capas_ui.append({'qty': c['qty'], 'cost': c['cost'], 'sub': sub})
        capas = capas_ui
    except Exception as e:
        messages.error(request, str(e))
        existencia = costo_total = costo_promedio = 0
        capas = []

    return render(request, 'inventario/reporte.html', {
        'producto': producto, 'metodo': metodo.upper(),
        'existencia': existencia, 'costo_total': costo_total,
        'costo_promedio': costo_promedio, 'capas': capas,
        'entrada_form': entrada_form, 'salida_form': salida_form
    })

# -------- Empleados --------
def employee_list(request):
    items = Employee.objects.order_by("nombre")
    form = EmployeeForm()  # formulario para usar en el modal
    return render(request, 'empleados/list.html', {'items': items, 'form': form})

def employee_create(request):
    form = EmployeeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Empleado creado.'); return redirect('employee_list')
    return render(request, 'empleados/form.html', {'form': form, 'titulo': 'Nuevo empleado'})

def employee_edit(request, pk):
    obj = get_object_or_404(Employee, pk=pk)
    form = EmployeeForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Empleado actualizado.'); return redirect('employee_list')
    return render(request, 'empleados/form.html', {'form': form, 'titulo': 'Editar empleado'})

@require_http_methods(['POST'])
def employee_delete(request, pk):
    get_object_or_404(Employee, pk=pk).delete()
    messages.info(request, 'Empleado eliminado.')
    return redirect('employee_list')

# -------- Parámetros y renta --------
def payroll_config(request):
    cfg = _get_config()
    form = PayrollConfigForm(request.POST or None, instance=cfg)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Parámetros guardados.')
        return redirect('payroll_config')
    return render(request, 'parametros/config.html', {'form': form})

def tax_list(request):
    return render(request, 'parametros/tax_list.html', {'items': TaxBracket.objects.order_by('desde')})

def tax_create(request):
    form = TaxBracketForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save(); messages.success(request, 'Tramo agregado.'); return redirect('tax_list')
    return render(request, 'parametros/tax_form.html', {'form': form})

@require_http_methods(['POST'])
def tax_delete(request, pk):
    get_object_or_404(TaxBracket, pk=pk).delete()
    messages.info(request, 'Tramo eliminado.')
    return redirect('tax_list')

# -------- Planilla --------
def planilla_periodo(request, anio, mes):
    """
    Muestra y (re)genera la planilla para el periodo (anio, mes).
    - GET: muestra la planilla (y la genera si no existe).
    - POST: regenera (recalcula) la planilla y vuelve a mostrarla.
    """
    # obtener/crear objeto periodo
    periodo, created = PayrollPeriod.objects.get_or_create(anio=anio, mes=mes)

    # Si viene POST -> forzar regeneración (por ejemplo desde el botón "Generar/Recalcular")
    if request.method == 'POST':
        try:
            generar_planilla(periodo)  # actualizar/crear líneas
            messages.success(request, f"Planilla {periodo} (re)generada correctamente.")
            return redirect('planilla_periodo', anio=anio, mes=mes)
        except Exception as e:
            messages.error(request, f"Error al generar planilla: {e}")

    # Asegurarnos de que la planilla está generada al menos una vez
    # generar_planilla siempre hace update_or_create, así nos aseguramos de tener las líneas.
    try:
        generar_planilla(periodo)
    except Exception as e:
        messages.error(request, f"Error al generar planilla: {e}")

    # Recuperar líneas y formatear valores para la plantilla
    lineas = periodo.lineas.select_related('empleado').all().order_by('empleado__nombre')

    # Normalizar y asegurar que los campos monetarios estén con 2 decimales
    for l in lineas:
        # si por algún motivo son None (seguro no lo serán), poner 0.00
        l.salario_base = (l.salario_base or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        l.isss = (l.isss or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        l.afp = (l.afp or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        l.renta = (l.renta or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        l.otras_deducciones = (l.otras_deducciones or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        l.liquido = (l.liquido or Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    total_liquido = sum((l.liquido for l in lineas), Decimal('0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # PASAMOS también la configuración actual para mostrar parámetros (opcional)
    cfg = _get_config()

    return render(request, 'planilla/planilla.html', {
        'periodo': periodo,
        'lineas': lineas,
        'total_liquido': total_liquido,
        'cfg': cfg,
    })
def planilla_excel(request, anio, mes):
    periodo, _ = PayrollPeriod.objects.get_or_create(anio=anio, mes=mes)
    lineas = generar_planilla(periodo)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Planilla {periodo}"
    headers = ["Empleado","Salario","ISSS","AFP","Renta","Líquido"]
    ws.append(headers)
    for l in lineas:
        ws.append([l.empleado.nombre, float(l.salario_base), float(l.isss), float(l.afp), float(l.renta), float(l.liquido)])
    ws.append(["", "", "", "", "Total", float(sum(l.liquido for l in lineas))])

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=planilla_{periodo.anio}_{periodo.mes}.xlsx'
    return resp

def planilla_pdf(request, anio, mes):
    periodo, _ = PayrollPeriod.objects.get_or_create(anio=anio, mes=mes)
    lineas = generar_planilla(periodo)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title=f"Planilla {periodo}")
    styles = getSampleStyleSheet()
    elems = [Paragraph(f"Planilla {periodo.anio}-{periodo.mes:02d}", styles['Title']), Spacer(1, 12)]

    data = [["Empleado","Salario","ISSS","AFP","Renta","Líquido"]]
    for l in lineas:
        data.append([l.empleado.nombre, f"${l.salario_base}", f"${l.isss}", f"${l.afp}", f"${l.renta}", f"${l.liquido}"])
    data.append(["","","","","Total", f"${sum(l.liquido for l in lineas)}"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), colors.black),
        ('TEXTCOLOR',(0,0),(-1,0), colors.whitesmoke),
        ('GRID',(0,0),(-1,-1), 0.5, colors.grey),
        ('ALIGN',(1,1),(-1,-1),'RIGHT'),
        ('FONTNAME',(0,0),(-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.whitesmoke, colors.Color(0.96,0.96,0.96)]),
    ]))
    elems.append(table)
    doc.build(elems)

    pdf = buffer.getvalue(); buffer.close()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename=planilla_{periodo.anio}_{periodo.mes}.pdf'
    return resp

def inventario_excel(request, producto_id, metodo):
    producto = get_object_or_404(Product, pk=producto_id)
    try:
        existencia, costo_total, costo_promedio, capas = valuacion_inventario(producto, metodo.upper())
    except ValueError as e:
        wb = Workbook(); ws = wb.active; ws.title = "Error"
        ws.append(["Error al generar inventario"])
        ws.append(["Producto", str(producto)])
        ws.append(["Método", metodo.upper()])
        ws.append([])
        ws.append([str(e)])
        ws.append(["Sugerencia", "Revise que las salidas no superen las existencias o ingrese entradas previas."])
        bio = BytesIO(); wb.save(bio); bio.seek(0)
        resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename=error_inventario_{producto.id}_{metodo}.xlsx'
        return resp

    wb = Workbook(); ws = wb.active; ws.title = f"{producto.codigo}-{metodo}"
    ws.append(["Producto", str(producto)])
    ws.append(["Método", metodo.upper()])
    ws.append(["Existencia", float(existencia)])
    ws.append(["Costo total", float(costo_total)])
    ws.append(["Costo promedio", float(costo_promedio)])
    ws.append([]); ws.append(["Capas","Cantidad","Costo unitario","Subtotal"])
    for c in capas:
        subtotal = float(c['qty'] * c['cost'])
        ws.append(["", float(c['qty']), float(c['cost']), subtotal])

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=inventario_{producto.codigo}_{metodo}.xlsx'
    return resp

def inventario_pdf(request, producto_id, metodo):
    producto = get_object_or_404(Product, pk=producto_id)
    try:
        existencia, costo_total, costo_promedio, capas = valuacion_inventario(producto, metodo.upper())
    except ValueError as e:
        # Generar un PDF simple con el error
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=LETTER, title=f"Error Inventario {producto} {metodo}")
        styles = getSampleStyleSheet()
        elems = [
            Paragraph("Error al generar inventario", styles['Title']),
            Spacer(1, 12),
            Paragraph(f"Producto: {producto}", styles['Normal']),
            Paragraph(f"Método: {metodo.upper()}", styles['Normal']),
            Spacer(1, 12),
            Paragraph(str(e), styles['Normal']),
            Spacer(1, 12),
            Paragraph("Sugerencia: revise que las salidas no superen las existencias o ingrese entradas previas.", styles['Italic']),
        ]
        doc.build(elems)
        pdf = buffer.getvalue(); buffer.close()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename=error_inventario_{producto.id}_{metodo}.pdf'
        return resp

    # ... (lo que ya tenías para el PDF cuando no hay error)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title=f"Inventario {producto} {metodo}")
    styles = getSampleStyleSheet()
    elems = [
        Paragraph(f"Inventario: {producto}", styles['Title']),
        Paragraph(f"Método: {metodo.upper()}", styles['Heading3']),
        Paragraph(f"Existencia: {existencia} — Costo total: ${costo_total} — Promedio: ${costo_promedio}", styles['Normal']),
        Spacer(1, 12)
    ]
    data = [["Cantidad","Costo unitario","Subtotal"]]
    for c in capas:
        subtotal = c['qty'] * c['cost']
        data.append([f"{c['qty']}", f"${c['cost']}", f"${subtotal}"])
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), colors.black),
        ('TEXTCOLOR',(0,0),(-1,0), colors.whitesmoke),
        ('GRID',(0,0),(-1,-1), 0.5, colors.grey),
        ('ALIGN',(0,1),(-1,-1),'RIGHT'),
        ('FONTNAME',(0,0),(-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.whitesmoke, colors.Color(0.96,0.96,0.96)]),
    ]))
    elems.append(table)
    doc.build(elems)

    pdf = buffer.getvalue(); buffer.close()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename=inventario_{producto.codigo}_{metodo}.pdf'
    return resp

def productos_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = ["Código", "Nombre", "Unidad"]
    ws.append(headers)

    # Estilos encabezado
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="000000")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = white
        cell.alignment = center
        cell.fill = fill
        cell.border = border

    # Datos
    qs = Product.objects.order_by("codigo").values_list("codigo", "nombre", "unidad")
    for r in qs:
        ws.append(list(r))

    # Bordes y auto ancho
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for c in row:
            c.border = border

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename=productos.xlsx'
    return resp


# ===== CSV =====
def productos_csv(request):
    import csv
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename=productos.csv'

    # BOM para abrir bien en Excel
    resp.write('\ufeff')

    writer = csv.writer(resp)
    writer.writerow(["Código", "Nombre", "Unidad"])
    for p in Product.objects.order_by("codigo"):
        writer.writerow([p.codigo, p.nombre, p.unidad])

    return resp


# ===== PDF =====
def productos_pdf(request):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(LETTER),
        title="Catálogo de Productos"
    )
    styles = getSampleStyleSheet()
    elems = [
        Paragraph("Catálogo de Productos", styles["Title"]),
        Spacer(1, 10)
    ]

    data = [["Código", "Nombre", "Unidad"]]
    qs = Product.objects.order_by("codigo").values_list("codigo", "nombre", "unidad")
    for row in qs:
        data.append(list(row))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.black),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.whitesmoke),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.Color(0.96,0.96,0.96)]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))
    elems.append(table)

    doc.build(elems)
    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename=productos.pdf'
    return resp

def importar_productos(request):
    import csv
    from io import TextIOWrapper
    from django.shortcuts import redirect, render
    from django.contrib import messages
    from django.db import transaction
    from .models import Product

    if request.method == "POST":
        file = request.FILES.get("archivo")
        if not file:
            messages.error(request, "⚠️ No se seleccionó ningún archivo.")
            return redirect("importar_productos")

        try:
            # UTF-8 con BOM (Excel-friendly)
            decoded = TextIOWrapper(file.file, encoding="utf-8-sig", newline="")
            sample = decoded.read(4096)
            decoded.seek(0)

            # Detectar delimitador , o ;
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            reader = csv.DictReader(decoded, dialect=dialect)

            # Normalizar encabezados a minúsculas
            if reader.fieldnames is None:
                messages.error(request, "⚠️ El archivo está vacío o sin encabezados.")
                return redirect("importar_productos")

            fieldnames = [f.strip().lower() for f in reader.fieldnames]
            required = {"codigo", "nombre", "unidad"}
            if not required.issubset(set(fieldnames)):
                messages.error(request, "⚠️ Encabezados requeridos: codigo,nombre,unidad.")
                return redirect("importar_productos")

            creados, actualizados = 0, 0
            with transaction.atomic():
                for row in reader:
                    # Re-map a minúsculas por si vienen mezclados
                    row = { (k or "").strip().lower(): (v or "").strip() for k, v in row.items() }

                    codigo = row.get("codigo", "")
                    nombre = row.get("nombre", "")
                    unidad = row.get("unidad", "unid") or "unid"

                    if not codigo or not nombre:
                        continue

                    obj, created = Product.objects.update_or_create(
                        codigo=codigo,
                        defaults={"nombre": nombre, "unidad": unidad},
                    )
                    creados += 1 if created else 0
                    actualizados += 0 if created else 1

            messages.success(request, f"✅ {creados} creados, {actualizados} actualizados.")
            return redirect("productos_list")

        except Exception as e:
            messages.error(request, f"⚠️ Archivo inválido: {e}")
            return redirect("importar_productos")

    return render(request, "productos/importar.html")

def productos_list(request):
    items = Product.objects.order_by("codigo")
    return render(request, "productos/list.html", {"items": items})

def productos_plantilla_csv(request):
    """
    Descarga una plantilla CSV vacía con encabezados: codigo,nombre,unidad
    (lista para completar y reimportar).
    """
    import csv
    from django.http import HttpResponse

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename=plantilla_productos.csv'

    # BOM para que Excel en español abra bien UTF-8
    resp.write('\ufeff')

    writer = csv.writer(resp)
    writer.writerow(["codigo", "nombre", "unidad"])
    # Si prefieres incluir ejemplos de referencia, descomenta estas líneas:
    # writer.writerow(["P001", "Tornillo 1/2\"", "unid"])
    # writer.writerow(["P002", "Tuerca 5mm", "unid"])

    return resp

def integrantes(request):
    integrantes = ["Manuel", "Johan", "Felipe"]
    return render(request, "integrantes.html", {"integrantes": integrantes})

@require_http_methods(['POST'])
def generar_planilla_now(request):
    """
    Crea/usa el periodo actual (año, mes) y (re)genera las líneas de planilla.
    Llama a generar_planilla(periodo) y muestra un mensaje.
    Redirige al home.
    """
    try:
        hoy = datetime.date.today()
        periodo, created = PayrollPeriod.objects.get_or_create(anio=hoy.year, mes=hoy.month)
        generar_planilla(periodo)
        if created:
            messages.success(request, f"Planilla {periodo} generada (nuevo periodo).")
        else:
            messages.success(request, f"Planilla {periodo} regenerada correctamente.")
    except Exception as e:
        messages.error(request, f"Error al generar planilla: {e}")
    return redirect('home')